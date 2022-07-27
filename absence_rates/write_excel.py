import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
import logging

logger = logging.getLogger(__name__)


class TagNotFound(Exception):
    """
    Exception raised for excel template not containing tags.

    Inputs:
        tag: the tag contained in the Excel template.
        message: the message printed when exception is raised.
        sheet: the Excel sheet missing the tag.
    
    Output:
        Raises an exception if tag is missing from Excel table.
    """

    def __init__(self, tag, sheet): # add sheet
        self.tag = tag
        self.sheet = sheet
        self.message = f"{tag} not found in sheet: {sheet}. Please amend the Excel template."
        super().__init__(self.message)


def find_starting_cell(wb, sheet, tag):
    """
    Returns the row and cell of the missing tag.

    Inputs:
        wb: the workbook sheet.
        sheet: the sheet's name. 
        tag: the tag in the Excel table template.
    
    Output:
        A tuple containing the tag's cell coordinates.
    """

    ws = wb[sheet]

    for row in ws.iter_rows(max_row = 1000):
        for cell in row:
            if cell.value == tag:
                return coordinate_to_tuple(cell.coordinate)
    else:
        raise TagNotFound(tag, sheet)


def prepare_table_1(csv_path):
    """
    Creates a function to prepare the data for Table 1 of the NHS Sickness Absence monthly tables.
    Orders the regions in the order they need to appear to mirror Table 1 in the Excel template.

    Inputs:
        csv_path: the location where the data is being pulled from.
        This is from the CSV and pasted into a dataframe named df.

    Output:
        A dataframe containing data for Table 1 of the NHS Sickness Absence monthly tables. 
    """
    logger.info(f"Preparing data for excel table 1")
    # This list sets out the order in which we want the values to appear
    order = ['ALL_ENGLAND',
            'London',
            'South West',
            'South East',
            'Midlands',
            'East of England',
            'North West',
            'North East and Yorkshire',
            'Special Health Authorities and other statutory bodies',]

    logger.info(f"Reading file from:\n{csv_path}")
    df = pd.read_csv(csv_path)
    df = df.round(2)
    df = df[(df['BREAKDOWN_TYPE'].isin(['ALL_ENGLAND', 'REGION']))]
    df['CONSTANT'] = 'ROW' # This column is only used to help us pivot the table onto one row
    df = df.pivot_table('SICKNESS_ABSENCE_RATE_PERCENT', 'CONSTANT', 'BREAKDOWN_VALUE')

    df = df[order]

    return df


def prepare_table_2(csv_path):
    """
    Creates a function to prepare the data for Table 2 of the NHS Sickness Absence monthly tables.
    Orders the staff groups in the order they need to appear to mirror Table 2 in the Excel template.

    Inputs:
        csv_path: the location where the data is being pulled from. 
        This is from the CSV and pasted into a dataframe named df.

    Output:
        A dataframe containing data for Table 2 of the NHS Sickness Absence monthly tables. 
    """
    logger.info(f"Preparing data for excel table 2")
    order = ['ALL_ENGLAND', 'Professionally qualified clinical staff',
    	'HCHS Doctors',	'Consultant',
        'Associate Specialist',
        'Specialty Doctor',
        'Staff Grade',
        'Specialty Registrar',
        'Core Training',
        'Foundation Doctor Year 2',
        'Foundation Doctor Year 1', 'Hospital Practitioner / Clinical Assistant',
        'Other and Local HCHS Doctor Grades',
        'Nurses & health visitors',
        'Midwives',
        'Ambulance staff',
        'Scientific, therapeutic & technical staff',
        'Support to clinical staff',
        'Support to doctors, nurses & midwives',
        'Support to ambulance staff',
        'Support to ST&T staff',
        'NHS infrastructure support',
        'Central functions',
        'Hotel, property & estates',
        'Senior managers',
        'Managers',
        'Other staff or those with unknown classification',
    ]

    logger.info(f"Reading file from:\n{csv_path}")
    df = pd.read_csv(csv_path)
    df = df.round(2)
    df = df[(df['BREAKDOWN_TYPE'].isin(['ALL_ENGLAND', 'MAJOR_STAFF_GROUPS', 'MINOR_STAFF_GROUPS', 'MINOR_STAFF_GRADES', ]))]
    df['CONSTANT'] = 'ROW' # This column is only used to help us pivot the table onto one row
    df = df.pivot_table('SICKNESS_ABSENCE_RATE_PERCENT', 'CONSTANT', 'BREAKDOWN_VALUE')
    
    df = df[order]

    return df


def prepare_table_2_1(df):
    """
    Creates a function to prepare the data for Table 2 of the NHS Sickness Absence monthly tables.
    Writing to Table 2 using a series of tags e.g. 2_1.
    
    Inputs:
        df: the dataframe where the data is being pulled from. 

    Output:
        A dataframe containing data for Table 2_1 tag of the NHS Sickness Absence monthly tables.
        This tag populates the All staff groups row. 
    """
    logger.info(f"Preparing data for excel tag table 2_1")

    logger.info(f"Reading df:\n{df}")
    df = df[["ALL_ENGLAND"]]

    return df


def prepare_table_2_2(df):
    """
    Creates a function to prepare the data for Table 2 of the NHS Sickness Absence monthly tables.
    Writing to Table 2 using a series of tags e.g. 2_1.
    
    Inputs:
        df: the dataframe where the data is being pulled from. 

    Output:
        A dataframe containing data for Table 2_2 tag of the NHS Sickness Absence monthly tables.
        This tag populates the Professionally qualified clinical staff row. 
    """
    logger.info(f"Preparing data for excel tag table 2_2")

    logger.info(f"Reading df:\n{df}")
    df = df[['Professionally qualified clinical staff']]

    return df


def prepare_table_2_3(df):
    """
    Creates a function to prepare the data for Table 2 of the NHS Sickness Absence monthly tables.
    Writing to Table 2 using a series of tags e.g. 2_1.
    
    Inputs:
        df: the dataframe where the data is being pulled from.

    Output:
        A dataframe containing data for Table 2_3 tag of the NHS Sickness Absence monthly tables.
        This tag populates the HCHS doctors row and the HCHS doctor grades. 
    """
    logger.info(f"Preparing data for excel tag table 2_3")

    logger.info(f"Reading df:\n{df}")
    df = df[['HCHS Doctors', 'Consultant',
        'Associate Specialist',
        'Specialty Doctor',
        'Staff Grade',
        'Specialty Registrar',
        'Core Training',
        'Foundation Doctor Year 2',
        'Foundation Doctor Year 1', 'Hospital Practitioner / Clinical Assistant',
        'Other and Local HCHS Doctor Grades']]
    df_t = df.transpose()

    return df_t


def prepare_table_2_4(df):
    """
    Creates a function to prepare the data for Table 2 of the NHS Sickness Absence monthly tables.
    Writing to Table 2 using a series of tags e.g. 2_1.
    
    Inputs:
        df: the dataframe where the data is being pulled from.

    Output:
        A dataframe containing data for Table 2_4 tag of the NHS Sickness Absence monthly tables.
        This tag populates the Nurses & HVs, Midwives, Ambulance staff & ST&T staff rows.
    """
    logger.info(f"Preparing data for excel tag table 2_4")

    logger.info(f"Reading df:\n{df}")
    df = df[['Nurses & health visitors',
        'Midwives',
        'Ambulance staff',
        'Scientific, therapeutic & technical staff']]
    df_t = df.transpose()

    return df_t


def prepare_table_2_5(df):
    """
    Creates a function to prepare the data for Table 2 of the NHS Sickness Absence monthly tables.
    Writing to Table 2 using a series of tags e.g. 2_1.
    
    Inputs:
        df: the dataframe where the data is being pulled from.

    Output:
        A dataframe containing data for Table 2_5 tag of the NHS Sickness Absence monthly tables. 
        This tag populates the Support to clinical staff rows.
    """
    logger.info(f"Preparing data for excel tag table 2_5")

    logger.info(f"Reading df:\n{df}")
    df = df[['Support to clinical staff',
        'Support to doctors, nurses & midwives',
        'Support to ambulance staff',
        'Support to ST&T staff',]]
    df_t = df.transpose()

    return df_t


def prepare_table_2_6(df):
    """
    Creates a function to prepare the data for Table 2 of the NHS Sickness Absence monthly tables.
    Writing to Table 2 using a series of tags e.g. 2_1.
    
    Inputs:
        df: the dataframe where the data is being pulled from.

    Output:
        A dataframe containing data for Table 2_6 tag of the NHS Sickness Absence monthly tables.
        This tag populates the NHS infrastructure support staff rows. 
    """
    logger.info(f"Preparing data for excel tag table 2_6")

    logger.info(f"Reading df:\n{df}")
    df = df[['NHS infrastructure support',
        'Central functions',
        'Hotel, property & estates',
        'Senior managers',
        'Managers']]
    df_t = df.transpose()

    return df_t



def prepare_table_2_7(df):
    """
    Creates a function to prepare the data for Table 2 of the NHS Sickness Absence monthly tables.
    Writing to Table 2 using a series of tags e.g. 2_1.
    
    Inputs:
        df: the dataframe where the data is being pulled from.

    Output:
        A dataframe containing data for Table 2_7 tag of the NHS Sickness Absence monthly tables.
        This tag populates the Unknown staff or those with unknown classification row. 
    """
    logger.info(f"Preparing data for excel tag table 2_7")

    logger.info(f"Reading df:\n{df}")
    df = df[["Other staff or those with unknown classification"]]

    return df


def prepare_table_3(csv_path):
    """
    Creates a function to prepare the data for Table 3 of the NHS Sickness Absence monthly tables.
    Orders the cluster groups in the order they need to appear to mirror Table 3 in the Excel template.

    Inputs:
        csv_path: the location where the data is being pulled from. 
        This is from the CSV and pasted into a dataframe named df.

    Output:
        A dataframe containing data for Table 3 of the NHS Sickness Absence monthly tables.
    """
    logger.info(f"Preparing data for excel table 3")

    # This list sets out the order in which we want the values to appear
    order = ['ALL_ENGLAND',
            'Acute',
            'Ambulance',
            'Clinical Commissioning Group',
            'Commissioning Support Unit',
            'Community Provider Trust',
            'Mental Health',
            'Special Health Authority',
            'Others',]

    logger.info(f"Reading file from:\n{csv_path}")
    df = pd.read_csv(csv_path)
    df = df.round(2)
    df = df[(df['BREAKDOWN_TYPE'].isin(['ALL_ENGLAND', 'ORGANISATION_TYPE']))]
    df['CONSTANT'] = 'ROW' # This column is only used to help us pivot the table onto one row
    df = df.pivot_table('SICKNESS_ABSENCE_RATE_PERCENT', 'CONSTANT', 'BREAKDOWN_VALUE')
    df = df.replace({9999: "."})

    df = df[order]
    return df


def write_tables_to_excel(tables, excel_template, excel_output):
    """
    Creates a function to write data to an excel template.

    Inputs:
        tables: the source of the data to be put into the templates.
        excel_template: the excel template tables which are in the repository.
        excel_output: where the populated tables are saved.

    Example:
        -------
        write_table_5_4_to_excel(table_data, excel_file=EXCEL_TEMPLATE)
    """
    logger.info(f"Writing tables to excel")
    logger.info(f"Using excel template:\n {excel_template}")
    xl_writer = pd.ExcelWriter(excel_output, engine='openpyxl')
    wb = load_workbook(excel_template)
    xl_writer.book = wb
    xl_writer.sheets = {ws.title: ws for ws in wb.worksheets}

    for table in tables:
        start_cell = find_starting_cell(wb, table["sheet_name"], table["tag"])
        table["data"].to_excel(
            xl_writer,
            table["sheet_name"],
            index=False,
            header=False,
            startcol=start_cell[1]-1,
            startrow=start_cell[0]-1
        )
    logger.info(f"Saving outputs to:\n {excel_output}")
    xl_writer.save()
